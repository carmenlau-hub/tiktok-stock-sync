"""
TikTok Stock Sync Tool — Streamlit app (single file)
Mister Mobile Singapore

Everything lives in this one file on purpose: with no second module there is
nothing that can fall out of sync when re-uploading to GitHub.

BUILD: 2026-08-02c
"""

from __future__ import annotations


import io
import re
from dataclasses import dataclass, field
from datetime import datetime
from difflib import SequenceMatcher

import openpyxl
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# --------------------------------------------------------------------------
# Mister Mobile brand palette (sampled from mistermobile.com.sg)
# --------------------------------------------------------------------------

MM_BLACK = "111111"    # nav bar / table headers
MM_YELLOW = "FFEB00"   # masthead / accent
MM_INK = "333333"      # body text
WHITE = "FFFFFF"
LIGHT = "FFFCE6"       # pale yellow tint
AMBER = "FFF7CC"       # editable "please fill this in" columns
GREEN = "E2EFDA"
RED = "FCE4E4"

NAVY = MM_BLACK  # backwards-compatible alias

REGISTRY_SHEETS = [
    "Locked Matches",
    "New Masterlist SKUs",
    "Match Review",
    "Not Selling in TikTok",
    "Not on TikTok Yet",
]

# Hidden lookup sheet backing the column-H dropdown / warning.
SKU_LOOKUP_SHEET = "_TikTokSKUs"

# TikTok "Template" sheet layout (fixed by TikTok Seller Center export)
TT_HEADER_KEY_ROW = 1      # product_id / sku_id / quantity ...
TT_LABEL_ROW = 3           # "Product ID" / "SKU ID" / "Quantity" ...
TT_FIRST_DATA_ROW = 6
TT_REQUIRED_KEYS = ["product_id", "product_name", "sku_id", "quantity"]

# Reviewer Decision dropdown — worded to mirror the Shopee registry.
DECISION_LINK = "Linked (fill col H)"
DECISION_NOT_SELLING = "Not Selling in TikTok"
DECISION_NOT_ON_TIKTOK = "Not on TikTok yet"
DECISION_PENDING = ""
DECISIONS = [DECISION_PENDING, DECISION_LINK, DECISION_NOT_SELLING, DECISION_NOT_ON_TIKTOK]
DECISION_CHOICES = [DECISION_LINK, DECISION_NOT_SELLING, DECISION_NOT_ON_TIKTOK]


class SyncError(Exception):
    """Raised for critical, export-blocking problems."""


# --------------------------------------------------------------------------
# Helpers
# --------------------------------------------------------------------------

def s(v) -> str:
    """Normalise any cell value to a clean string ('' for blanks)."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).replace("\xa0", " ").strip()


def to_int(v, default=0) -> int:
    if v is None or v == "":
        return default
    try:
        return int(float(str(v).replace(",", "").strip()))
    except (TypeError, ValueError):
        return default


_norm_re = re.compile(r"[^a-z0-9]+")


def norm(text: str) -> str:
    """Lowercase alphanumeric token string used for fuzzy comparison."""
    return _norm_re.sub(" ", s(text).lower()).strip()


def tokens(text: str) -> set:
    return {t for t in norm(text).split() if t}


def similarity(a: str, b: str) -> float:
    """Blend of token overlap, coverage and sequence ratio, 0..1."""
    na, nb = norm(a), norm(b)
    if not na or not nb:
        return 0.0
    ta, tb = set(na.split()), set(nb.split())
    inter = len(ta & tb)
    jac = inter / len(ta | tb) if (ta | tb) else 0.0
    cover = inter / min(len(ta), len(tb)) if ta and tb else 0.0
    seq = SequenceMatcher(None, na, nb).ratio()
    return 0.40 * jac + 0.35 * cover + 0.25 * seq


def split_ids(v) -> list:
    """'23300, 30639' or '23300|30639' -> ['23300', '30639']"""
    raw = s(v)
    if not raw:
        return []
    return [p.strip() for p in re.split(r"[,;|/\s]+", raw) if p.strip()]


def classify_decision(text: str) -> str:
    """Map any decision wording (old or new) onto a canonical constant."""
    n = norm(text)
    if not n:
        return DECISION_PENDING
    if "not selling" in n:
        return DECISION_NOT_SELLING
    if "not on" in n:
        return DECISION_NOT_ON_TIKTOK
    if any(k in n for k in ("link", "confirm", "lock")):
        return DECISION_LINK
    return DECISION_PENDING


# --------------------------------------------------------------------------
# Data models
# --------------------------------------------------------------------------

@dataclass
class PosItem:
    stock_id: str
    category: str = ""
    brand: str = ""
    model: str = ""
    color: str = ""
    available: int = 0

    @property
    def label(self) -> str:
        bits = [b for b in (self.brand, self.model, self.color) if b]
        return " ".join(bits) or self.stock_id

    @property
    def search_text(self) -> str:
        return " ".join(x for x in (self.category, self.brand, self.model, self.color) if x)


@dataclass
class TtRow:
    row: int                 # 1-based row index in the Template sheet
    product_id: str = ""
    category: str = ""
    product_name: str = ""
    sku_id: str = ""
    variation_value: str = ""
    price: str = ""
    quantity: int = 0
    seller_sku: str = ""

    @property
    def search_text(self) -> str:
        return f"{self.seller_sku} {self.variation_value} {self.product_name}"

    @property
    def display(self) -> str:
        head = self.seller_sku or self.product_name
        return f"{head} — {self.variation_value}" if self.variation_value else head


@dataclass
class ParsedRegistry:
    locked: dict = field(default_factory=dict)        # tt_sku_id -> [pos_id, ...]
    match_review: set = field(default_factory=set)    # tt_sku_id
    not_selling: set = field(default_factory=set)     # pos_id
    not_on_tiktok: set = field(default_factory=set)   # pos_id
    new_decisions: dict = field(default_factory=dict) # pos_id -> (decision, tt_sku_id, notes)
    sheets_found: list = field(default_factory=list)
    warnings: list = field(default_factory=list)


# --------------------------------------------------------------------------
# POS Masterlist parsing
# --------------------------------------------------------------------------

POS_FIELD_ALIASES = {
    "stock_id": ["stock type id", "stock id", "masterlist stock type id", "stocktypeid",
                 "pos stock type id"],
    "category": ["category"],
    "brand": ["brand"],
    "model": ["model"],
    "color": ["color", "colour"],
    "available": ["available quantity", "available qty", "availableqty", "total"],
}


def _match_alias(header: str, aliases: list) -> bool:
    h = norm(header)
    return any(h == norm(a) for a in aliases)


def parse_pos_masterlist(file_bytes: bytes) -> tuple:
    """
    Returns (dict[pos_id -> PosItem], warnings list).

    Handles the POS stock report layout where row 1 holds group headers
    (Stock Type ID / Category / Brand / Model / Color / Total / <branches>)
    and row 2 holds sub-headers ("Available Quantity", "Reserved Quantity"...).
    Available Qty is taken from the FIRST column whose row-2 label is
    'Available Quantity' — the company-wide Total block.
    """
    warnings: list = []
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = []
    for r in ws.iter_rows(values_only=True):
        rows.append(list(r))
        if len(rows) > 200000:
            break
    wb.close()

    if len(rows) < 3:
        raise SyncError("POS Masterlist appears to be empty.")

    header1 = [s(v) for v in rows[0]]
    header2 = [s(v) for v in rows[1]] if len(rows) > 1 else []

    colmap: dict = {}
    for idx, h in enumerate(header1):
        for field_name, aliases in POS_FIELD_ALIASES.items():
            if field_name == "available":
                continue
            if field_name not in colmap and _match_alias(h, aliases):
                colmap[field_name] = idx

    avail_idx = None
    for idx, h in enumerate(header2):
        if norm(h) in ("available quantity", "available qty"):
            avail_idx = idx
            break
    two_row_header = avail_idx is not None

    if avail_idx is None:
        for idx, h in enumerate(header1):
            if _match_alias(h, POS_FIELD_ALIASES["available"]):
                avail_idx = idx
                break

    if "stock_id" not in colmap:
        raise SyncError(
            "POS Masterlist is missing a 'Stock Type ID' column. "
            f"Found headers: {', '.join([h for h in header1 if h][:12])}"
        )
    if avail_idx is None:
        raise SyncError("POS Masterlist is missing an 'Available Quantity' / 'Total' column.")

    colmap["available"] = avail_idx
    first_data = 2 if two_row_header else 1

    items: dict = {}
    dupes = 0
    for r in rows[first_data:]:
        if not r:
            continue

        def get(name, default=""):
            i = colmap.get(name)
            if i is None or i >= len(r):
                return default
            return r[i]

        pid = s(get("stock_id"))
        if not pid:
            continue
        qty = to_int(get("available"))
        item = PosItem(
            stock_id=pid,
            category=s(get("category")),
            brand=s(get("brand")),
            model=s(get("model")),
            color=s(get("color")),
            available=qty,
        )
        if pid in items:
            dupes += 1
            items[pid].available += qty
        else:
            items[pid] = item

    if dupes:
        warnings.append(
            f"{dupes} duplicate Stock Type ID row(s) in POS Masterlist — quantities were summed."
        )
    if not items:
        raise SyncError("No POS Masterlist rows could be read.")

    return items, warnings


# --------------------------------------------------------------------------
# TikTok template parsing
# --------------------------------------------------------------------------

def parse_tiktok_template(file_bytes: bytes) -> tuple:
    """Returns (list[TtRow], meta dict). Raises SyncError on structural problems."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False)

    if "Template" not in wb.sheetnames:
        raise SyncError(
            "TikTok file has no worksheet named 'Template'. "
            f"Worksheets found: {', '.join(wb.sheetnames)}"
        )
    ws = wb["Template"]

    keys = {}
    for c in range(1, ws.max_column + 1):
        k = norm(ws.cell(TT_HEADER_KEY_ROW, c).value).replace(" ", "_")
        if k:
            keys[k] = c

    missing = [k for k in TT_REQUIRED_KEYS if k not in keys]
    if missing:
        raise SyncError(
            "TikTok Template sheet is missing required column(s): "
            + ", ".join(missing)
            + f". Detected: {', '.join(sorted(keys))}"
        )

    rows: list = []
    for r in range(TT_FIRST_DATA_ROW, ws.max_row + 1):
        sku = s(ws.cell(r, keys["sku_id"]).value)
        pid = s(ws.cell(r, keys["product_id"]).value)
        if not sku and not pid:
            continue
        rows.append(
            TtRow(
                row=r,
                product_id=pid,
                category=s(ws.cell(r, keys["category"]).value) if "category" in keys else "",
                product_name=s(ws.cell(r, keys["product_name"]).value),
                sku_id=sku,
                variation_value=s(ws.cell(r, keys["variation_value"]).value)
                if "variation_value" in keys else "",
                price=s(ws.cell(r, keys["price"]).value) if "price" in keys else "",
                quantity=to_int(ws.cell(r, keys["quantity"]).value),
                seller_sku=s(ws.cell(r, keys["seller_sku"]).value) if "seller_sku" in keys else "",
            )
        )

    wb.close()

    if not rows:
        raise SyncError("TikTok Template sheet contains no listing rows (expected data from row 6).")

    meta = {
        "quantity_col": keys["quantity"],
        "sku_col": keys["sku_id"],
        "columns": keys,
        "row_count": len(rows),
    }
    return rows, meta


# --------------------------------------------------------------------------
# SKU Registry parsing
# --------------------------------------------------------------------------

def _header_index(ws) -> dict:
    idx = {}
    for c in range(1, ws.max_column + 1):
        h = norm(ws.cell(1, c).value)
        if h:
            idx[h] = c
    return idx


def _find_col(idx: dict, *candidates):
    for cand in candidates:
        n = norm(cand)
        if n in idx:
            return idx[n]
    for cand in candidates:
        n = norm(cand)
        for k, v in idx.items():
            if n and n in k:
                return v
    return None


def parse_registry(file_bytes: bytes | None) -> ParsedRegistry:
    """Reads the 5 registry worksheets. Missing file -> empty registry."""
    reg = ParsedRegistry()
    if not file_bytes:
        return reg

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    names = {norm(n): n for n in wb.sheetnames}
    reg.sheets_found = list(wb.sheetnames)

    def sheet(*aliases):
        for a in aliases:
            if norm(a) in names:
                return wb[names[norm(a)]]
        return None

    # ---- Locked Matches -------------------------------------------------
    ws = sheet("Locked Matches", "Locked Match")
    if ws is None:
        reg.warnings.append("Worksheet 'Locked Matches' not found in the SKU file.")
    else:
        idx = _header_index(ws)
        c_sku = _find_col(idx, "sku id", "variation id", "tiktok sku id", "tiktok variation id")
        c_ml = _find_col(idx, "locked masterlist id s", "locked masterlist id", "masterlist id")
        if c_sku is None or c_ml is None:
            reg.warnings.append("'Locked Matches' is missing a SKU ID or LOCKED Masterlist ID(s) column.")
        else:
            for r in ws.iter_rows(min_row=2, values_only=True):
                sku = s(r[c_sku - 1]) if c_sku - 1 < len(r) else ""
                mls = split_ids(r[c_ml - 1]) if c_ml - 1 < len(r) else []
                if sku and mls:
                    reg.locked.setdefault(sku, [])
                    for m in mls:
                        if m not in reg.locked[sku]:
                            reg.locked[sku].append(m)

    # ---- Match Review ---------------------------------------------------
    ws = sheet("Match Review")
    if ws is None:
        reg.warnings.append("Worksheet 'Match Review' not found in the SKU file.")
    else:
        idx = _header_index(ws)
        c_sku = _find_col(idx, "sku id", "variation id", "tiktok sku id")
        c_fix = _find_col(idx, "corrected masterlist id", "masterlist id")
        c_dec = _find_col(idx, "reviewer decision", "decision")
        if c_sku is not None:
            for r in ws.iter_rows(min_row=2, values_only=True):
                sku = s(r[c_sku - 1]) if c_sku - 1 < len(r) else ""
                if not sku:
                    continue
                reg.match_review.add(sku)
                fix = split_ids(r[c_fix - 1]) if c_fix and c_fix - 1 < len(r) else []
                dec = s(r[c_dec - 1]) if c_dec and c_dec - 1 < len(r) else ""
                # A reviewed + corrected Match Review row is promoted to locked.
                if fix and classify_decision(dec) == DECISION_LINK:
                    reg.locked.setdefault(sku, [])
                    for m in fix:
                        if m not in reg.locked[sku]:
                            reg.locked[sku].append(m)
                    reg.match_review.discard(sku)

    # ---- Classification sheets -----------------------------------------
    for alias_set, target in (
        (("Not Selling in TikTok", "Not Selling in Shopee", "Not Selling in IShopChangi",
          "Not Selling"), reg.not_selling),
        (("Not on TikTok Yet", "Not on Shopee Yet", "Not on IShopChangi Yet",
          "Not on TikTok"), reg.not_on_tiktok),
    ):
        ws = sheet(*alias_set)
        if ws is None:
            reg.warnings.append(f"Worksheet '{alias_set[0]}' not found in the SKU file.")
            continue
        idx = _header_index(ws)
        c_id = _find_col(idx, "masterlist stock type id", "stock type id", "stock id")
        if c_id is None:
            reg.warnings.append(f"'{alias_set[0]}' is missing a Masterlist Stock Type ID column.")
            continue
        for r in ws.iter_rows(min_row=2, values_only=True):
            pid = s(r[c_id - 1]) if c_id - 1 < len(r) else ""
            if pid:
                target.add(pid)

    # ---- New Masterlist SKUs (previous decisions) -----------------------
    ws = sheet("New Masterlist SKUs", "New Masterlist SKU")
    if ws is None:
        reg.warnings.append("Worksheet 'New Masterlist SKUs' not found in the SKU file.")
    else:
        idx = _header_index(ws)
        c_id = _find_col(idx, "masterlist stock type id", "stock type id")
        c_link = _find_col(idx, "link to tiktok sku id", "link to tiktok variation id",
                           "link to shopee variation id", "link to sku id", "link")
        c_dec = _find_col(idx, "reviewer decision", "decision")
        c_note = _find_col(idx, "notes", "note")
        if c_id is not None:
            for r in ws.iter_rows(min_row=2, values_only=True):
                pid = s(r[c_id - 1]) if c_id - 1 < len(r) else ""
                if not pid:
                    continue
                link = s(r[c_link - 1]) if c_link and c_link - 1 < len(r) else ""
                dec = s(r[c_dec - 1]) if c_dec and c_dec - 1 < len(r) else ""
                note = s(r[c_note - 1]) if c_note and c_note - 1 < len(r) else ""
                if dec or link:
                    reg.new_decisions[pid] = (dec, link, note)

    wb.close()
    return reg


def parse_seed_map(file_bytes: bytes | None) -> dict:
    """Optional 'POS stock ID and TikTok Variation ID' sheet -> {tt_sku_id: [pos_id]}."""
    out: dict = {}
    if not file_bytes:
        return out
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    idx = {}
    first = True
    c_pos = c_tt = None
    for r in ws.iter_rows(values_only=True):
        if first:
            first = False
            for i, v in enumerate(r):
                idx[norm(v)] = i
            c_pos = next((i for k, i in idx.items() if "stock" in k and "id" in k), 0)
            c_tt = next((i for k, i in idx.items()
                         if ("variation" in k or "sku" in k) and "id" in k and i != c_pos), 1)
            continue
        pos = s(r[c_pos]) if c_pos < len(r) else ""
        tt = s(r[c_tt]) if c_tt < len(r) else ""
        if pos and tt:
            out.setdefault(tt, [])
            if pos not in out[tt]:
                out[tt].append(pos)
    wb.close()
    return out


# --------------------------------------------------------------------------
# Matching engine
# --------------------------------------------------------------------------

@dataclass
class SyncResult:
    locked_rows: list = field(default_factory=list)
    review_rows: list = field(default_factory=list)
    new_skus: list = field(default_factory=list)
    not_selling: list = field(default_factory=list)
    not_on_tiktok: list = field(default_factory=list)
    errors: list = field(default_factory=list)
    warnings: list = field(default_factory=list)
    zero_stock_notes: list = field(default_factory=list)
    qty_updates: dict = field(default_factory=dict)
    changed: int = 0
    unchanged: int = 0

    @property
    def stats(self) -> dict:
        return {
            "Locked Matches updated": len(self.locked_rows),
            "Quantity cells changed": self.changed,
            "New Masterlist SKUs found": len(self.new_skus),
            "SKUs requiring review": len(self.new_skus) + len(self.review_rows),
            "Not Selling in TikTok": len(self.not_selling),
            "Not on TikTok Yet": len(self.not_on_tiktok),
            "Validation errors": len(self.errors),
            "Unmatched / invalid records": len(self.review_rows),
        }


def build_sync(pos: dict, tt_rows: list, reg: ParsedRegistry, seed: dict | None = None,
               overrides: dict | None = None) -> SyncResult:
    """
    overrides: {pos_id: {"decision": str, "sku_id": str, "notes": str}} — the
    in-session review decisions made in the Streamlit UI.
    """
    res = SyncResult()
    overrides = overrides or {}
    seed = seed or {}

    locked: dict = {k: list(v) for k, v in reg.locked.items()}
    for sku, ids in seed.items():
        if sku not in locked:
            locked[sku] = list(ids)

    not_selling = set(reg.not_selling)
    not_on_tt = set(reg.not_on_tiktok)

    for pid, o in overrides.items():
        dec = classify_decision(o.get("decision", ""))
        if dec == DECISION_LINK and o.get("sku_id"):
            sku = s(o["sku_id"])
            locked.setdefault(sku, [])
            if pid not in locked[sku]:
                locked[sku].append(pid)
            not_selling.discard(pid)
            not_on_tt.discard(pid)
        elif dec == DECISION_NOT_SELLING:
            not_selling.add(pid)
        elif dec == DECISION_NOT_ON_TIKTOK:
            not_on_tt.add(pid)

    by_sku = {r.sku_id: r for r in tt_rows}

    # ---- Locked matches -> quantity updates -----------------------------
    linked_pos_ids: set = set()
    for sku, ids in locked.items():
        row = by_sku.get(sku)
        valid_ids = [p for p in ids if p in pos]
        missing_ids = [p for p in ids if p not in pos]

        if row is None:
            res.errors.append(
                f"Locked SKU ID {sku} is not present in the uploaded TikTok inventory file — skipped."
            )
            continue
        if missing_ids:
            # A Masterlist ID absent from today's stock report means the item has
            # no stock today -> it contributes 0, and the listing is still synced.
            res.zero_stock_notes.append(
                f"SKU ID {sku}: Masterlist ID(s) {', '.join(missing_ids)} absent from "
                "today's POS Masterlist — counted as 0."
            )

        linked_pos_ids.update(valid_ids)
        target = max(sum(pos[p].available for p in valid_ids), 0)

        res.locked_rows.append({
            "product_id": row.product_id,
            "product_name": row.product_name,
            "sku_id": row.sku_id,
            "variation_value": row.variation_value,
            "seller_sku": row.seller_sku,
            # ALL ids are written back, including ones absent from today's POS
            # report — otherwise the lock would be lost on the next run.
            "masterlist_ids": ", ".join(ids),
            "ml_labels": " | ".join(
                f"{p}:{pos[p].model}|{pos[p].color}" if p in pos else f"{p}:(not in POS today)"
                for p in ids
            ),
            "ml_qty": ", ".join(str(pos[p].available) if p in pos else "0" for p in ids),
            "current": row.quantity,
            "target": target,
            "n": len(ids),
        })

        if target != row.quantity:
            res.qty_updates[row.row] = target
            res.changed += 1
        else:
            res.unchanged += 1

    # ---- Match Review: every TikTok listing not locked -------------------
    for row in tt_rows:
        if row.sku_id in locked:
            continue
        res.review_rows.append({
            "product_id": row.product_id,
            "product_name": row.product_name,
            "sku_id": row.sku_id,
            "variation_value": row.variation_value,
            "seller_sku": row.seller_sku,
            "current": row.quantity,
        })

    # ---- New Masterlist SKUs --------------------------------------------
    for pid, item in pos.items():
        if item.available <= 0:
            continue
        if pid in linked_pos_ids or pid in not_selling or pid in not_on_tt:
            continue
        res.new_skus.append(item)
    res.new_skus.sort(key=lambda i: (-i.available, i.label))

    res.not_selling = [pos.get(p) or PosItem(stock_id=p) for p in sorted(not_selling, key=s)]
    res.not_on_tiktok = [pos.get(p) or PosItem(stock_id=p) for p in sorted(not_on_tt, key=s)]
    res.warnings.extend(reg.warnings)
    return res


class MatchIndex:
    """
    Pre-tokenised inverted index over the TikTok listings so that suggesting
    matches for one POS item touches only a few hundred candidate rows instead
    of all ~3,400. Build once per uploaded file and reuse.
    """

    _STOP = {"5g", "4g", "gb", "tb", "mm", "new", "used", "sg", "set", "local",
             "phone", "case", "the", "and", "with", "for"}

    def __init__(self, tt_rows: list):
        self.rows = tt_rows
        self.seller, self.full, self.var, self.toks = [], [], [], []
        self.postings: dict = {}
        for i, r in enumerate(tt_rows):
            ns = norm(r.seller_sku)
            nf = norm(f"{r.product_name} {r.variation_value}")
            nv = norm(r.variation_value)
            self.seller.append(ns)
            self.full.append(nf)
            self.var.append(nv)
            tk = set(ns.split()) | set(nf.split())
            tk.discard("")
            self.toks.append(tk)
            for t in tk:
                self.postings.setdefault(t, []).append(i)
        limit = max(400, len(tt_rows) // 4)
        self.common = {t for t, v in self.postings.items() if len(v) > limit} | self._STOP

    def _score(self, qn: str, qt: set, i: int) -> float:
        best = 0.0
        for cand in (self.seller[i], self.full[i], self.var[i]):
            if not cand:
                continue
            ct = self.toks[i] if cand is not self.var[i] else set(cand.split())
            inter = len(qt & ct)
            if not inter:
                continue
            union = len(qt | ct)
            jac = inter / union if union else 0.0
            cover = inter / min(len(qt), len(ct)) if qt and ct else 0.0
            seq = SequenceMatcher(None, qn, cand).ratio()
            best = max(best, 0.40 * jac + 0.35 * cover + 0.25 * seq)
        return best

    def suggest(self, item: PosItem, taken: set | None = None, top: int = 6) -> list:
        taken = taken or set()
        qn = norm(item.search_text)
        qt = {t for t in qn.split() if t}
        if not qt:
            return []

        keys = sorted((t for t in qt if t not in self.common),
                      key=lambda t: len(self.postings.get(t, ())))
        cand: set = set()
        for t in keys:
            cand.update(self.postings.get(t, ()))
            if len(cand) > 1200:
                break
        if not cand:
            for t in qt:
                cand.update(self.postings.get(t, ())[:600])
        if not cand:
            return []

        scored = []
        for i in cand:
            if self.rows[i].sku_id in taken:
                continue
            sc = self._score(qn, qt, i)
            if sc > 0.20:
                scored.append((self.rows[i], sc))
        scored.sort(key=lambda x: -x[1])
        return scored[:top]


def suggest_matches(item: PosItem, tt_rows_or_index, taken: set | None = None, top: int = 6) -> list:
    """Accepts either a MatchIndex (fast) or a plain list of TtRow."""
    idx = (tt_rows_or_index if isinstance(tt_rows_or_index, MatchIndex)
           else MatchIndex(tt_rows_or_index))
    return idx.suggest(item, taken, top)


# --------------------------------------------------------------------------
# Exports
# --------------------------------------------------------------------------

def export_tiktok_template(original_bytes: bytes, qty_updates: dict, quantity_col: int) -> bytes:
    """
    Reopen the ORIGINAL TikTok workbook and write only the Quantity cells.
    Every other cell, sheet, style and structure is untouched.
    """
    wb = openpyxl.load_workbook(io.BytesIO(original_bytes))
    ws = wb["Template"]
    for row_idx, qty in qty_updates.items():
        ws.cell(row_idx, quantity_col).value = int(qty)
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


def _style_header(ws, ncols: int, widths: list | None = None):
    """Black header row with yellow text — the Mister Mobile nav bar."""
    fill = PatternFill("solid", fgColor=MM_BLACK)
    for c in range(1, ncols + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True, color=MM_YELLOW)
        cell.fill = fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{max(ws.max_row, 1)}"
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w


def _add_decision_dropdown(ws, col_letter: str, last_row: int):
    """Reviewer Decision dropdown, worded like the Shopee registry."""
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(DECISION_CHOICES) + '"',
        allow_blank=True,
        showDropDown=False,   # False = SHOW the in-cell arrow (Excel quirk)
    )
    dv.errorTitle = "Pick from the list"
    dv.error = "Choose one of: " + " / ".join(DECISION_CHOICES)
    dv.promptTitle = "Reviewer decision"
    dv.prompt = ("Linked (fill col H) — enter the TikTok SKU ID in column H.\n"
                 "Not Selling in TikTok — we stock it but don't list it.\n"
                 "Not on TikTok yet — no listing exists.")
    dv.showInputMessage = True
    dv.showErrorMessage = True
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}2:{col_letter}{max(last_row, 2)}")


def _add_sku_lookup(wb, sku_ids: list):
    """Hidden sheet holding every valid TikTok SKU ID, for column-H validation."""
    ws = wb.create_sheet(SKU_LOOKUP_SHEET)
    for i, sku in enumerate(sku_ids, start=1):
        ws.cell(i, 1, sku)
    ws.column_dimensions["A"].width = 22
    ws.sheet_state = "hidden"
    return len(sku_ids)


def _add_sku_validation(ws, col_letter: str, last_row: int, n_skus: int):
    """Warn (don't block) when column H holds an ID that isn't a real TikTok SKU."""
    if n_skus <= 0:
        return
    dv = DataValidation(
        type="list",
        formula1=f"={SKU_LOOKUP_SHEET}!$A$1:$A${n_skus}",
        allow_blank=True,
        showDropDown=True,    # too many entries to browse; validation only
    )
    dv.errorStyle = "warning"   # user can override and keep the value
    dv.errorTitle = "Not a known TikTok SKU ID"
    dv.error = ("That ID isn't in the TikTok inventory file you exported from. "
                "Check for a typo, or click Yes to keep it anyway.")
    dv.showErrorMessage = True
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}2:{col_letter}{max(last_row, 2)}")


def export_registry(res: SyncResult, pos: dict, applied_note: str = "",
                    sku_ids: list | None = None) -> bytes:
    """
    Build the SKU Registry workbook in Mister Mobile colours.

    sku_ids: every valid TikTok SKU ID. When supplied, column H of
    'New Masterlist SKUs' warns on IDs that don't exist.
    """
    wb = openpyxl.Workbook()
    amber = PatternFill("solid", fgColor=AMBER)

    # ---- Summary --------------------------------------------------------
    ws = wb.active
    ws.title = "Summary"
    for col in ("B", "C"):
        ws[f"{col}2"].fill = PatternFill("solid", fgColor=MM_YELLOW)
    ws["B2"] = "TikTok Stock Bulk Update — Match Registry"
    ws["B2"].font = Font(bold=True, size=15, color=MM_BLACK)
    ws["B3"] = f"Mister Mobile Singapore · generated {datetime.now().strftime('%d-%m-%Y %H:%M')}"
    ws["B3"].font = Font(size=9, italic=True, color=MM_INK)

    ws["B5"], ws["C5"] = "Category", "Count"
    for c in ("B5", "C5"):
        ws[c].font = Font(bold=True, color=MM_YELLOW)
        ws[c].fill = PatternFill("solid", fgColor=MM_BLACK)

    r = 6
    for k, v in res.stats.items():
        ws.cell(r, 2, k).font = Font(size=10)
        ws.cell(r, 3, v)
        r += 1
    if applied_note:
        ws.cell(r + 1, 2, applied_note).font = Font(size=9, italic=True)
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 12

    # ---- Locked Matches -------------------------------------------------
    ws = wb.create_sheet("Locked Matches")
    ws.append(["#", "TikTok Product ID", "TikTok Product Name", "SKU ID", "Variation Option",
               "Seller SKU", "LOCKED Masterlist ID(s)", "ML Model(s)|Color", "ML Available Qty",
               "Target Stock", "# SKUs"])
    for i, d in enumerate(res.locked_rows, 1):
        ws.append([i, d["product_id"], d["product_name"], d["sku_id"], d["variation_value"],
                   d["seller_sku"], d["masterlist_ids"], d["ml_labels"], d["ml_qty"],
                   d["target"], d["n"]])
    _style_header(ws, 11, [5, 20, 34, 20, 26, 30, 20, 34, 14, 11, 7])

    # ---- New Masterlist SKUs -------------------------------------------
    ws = wb.create_sheet("New Masterlist SKUs")
    ws.append(["#", "Masterlist Stock Type ID", "Category", "Brand", "Model", "Color",
               "Available Qty", "Link to TikTok SKU ID", "Reviewer Decision", "Notes"])
    for i, it in enumerate(res.new_skus, 1):
        ws.append([i, it.stock_id, it.category, it.brand, it.model, it.color,
                   it.available, "", "", ""])
    _style_header(ws, 10, [5, 22, 12, 14, 30, 18, 13, 24, 24, 26])
    last = ws.max_row
    for row in ws.iter_rows(min_row=2, min_col=8, max_col=10):
        for cell in row:
            cell.fill = amber
    _add_decision_dropdown(ws, "I", last)
    new_sku_sheet_last = last

    # ---- Match Review ---------------------------------------------------
    ws = wb.create_sheet("Match Review")
    ws.append(["#", "TikTok Product ID", "TikTok Product Name", "SKU ID", "Variation Option",
               "Seller SKU", "Current Seller Stock", "Corrected Masterlist ID",
               "Reviewer Decision", "Notes"])
    for i, d in enumerate(res.review_rows, 1):
        ws.append([i, d["product_id"], d["product_name"], d["sku_id"], d["variation_value"],
                   d["seller_sku"], d["current"], "", "", ""])
    _style_header(ws, 10, [5, 20, 34, 20, 26, 30, 16, 22, 22, 26])
    for row in ws.iter_rows(min_row=2, min_col=8, max_col=10):
        for cell in row:
            cell.fill = amber

    # ---- Classification sheets -----------------------------------------
    for title, items in (("Not Selling in TikTok", res.not_selling),
                         ("Not on TikTok Yet", res.not_on_tiktok)):
        ws = wb.create_sheet(title)
        ws.append(["#", "Masterlist Stock Type ID", "Category", "Brand", "Model", "Color",
                   "Available Qty"])
        for i, it in enumerate(items, 1):
            ws.append([i, it.stock_id, it.category, it.brand, it.model, it.color, it.available])
        _style_header(ws, 7, [5, 22, 12, 14, 30, 18, 13])

    # ---- Validation Errors ---------------------------------------------
    ws = wb.create_sheet("Validation Errors")
    ws.append(["#", "Severity", "Issue"])
    i = 0
    for sev, bucket in (("Error", res.errors), ("Warning", res.warnings),
                        ("Info", res.zero_stock_notes)):
        for e in bucket:
            i += 1
            ws.append([i, sev, e])
    _style_header(ws, 3, [5, 12, 110])

    # ---- Hidden SKU lookup + column H validation ------------------------
    if sku_ids:
        n = _add_sku_lookup(wb, sku_ids)
        _add_sku_validation(wb["New Masterlist SKUs"], "H", new_sku_sheet_last, n)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ==========================================================================
# Streamlit interface
# ==========================================================================

st.set_page_config(page_title="TikTok Stock Sync Tool", page_icon="📦", layout="wide")

st.markdown(
    """
    <style>
      /* ---- Mister Mobile brand palette ------------------------------- */
      :root {
        --mm-yellow: #FFEB00;
        --mm-black: #111111;
        --mm-ink: #333333;
        --mm-line: #E6E6E6;
        --mm-cream: #FFFDF0;
      }

      .block-container {padding-top: 1.6rem; max-width: 1400px;}

      /* Yellow masthead + black nav strip, mirroring mistermobile.com.sg */
      .mm-hero {background: var(--mm-yellow); padding: 1.05rem 1.4rem;
                border-radius: 10px 10px 0 0; color: var(--mm-black);}
      .mm-hero h1 {margin:0; font-size:1.55rem; color:var(--mm-black);
                   font-weight:800; letter-spacing:-.01em;}
      .mm-bar {background: var(--mm-black); color:#fff; padding:.5rem 1.4rem;
               border-radius: 0 0 10px 10px; margin-bottom:1.3rem;
               font-size:.85rem; letter-spacing:.02em;}
      .mm-bar b {color: var(--mm-yellow); font-weight:600;}

      .mm-sku {font-family: ui-monospace, Menlo, monospace; font-size:.8rem; color:#6B6B6B;}

      div[data-testid="stMetric"] {background: var(--mm-cream);
        border:1px solid var(--mm-line); border-left:4px solid var(--mm-yellow);
        border-radius:6px; padding:.65rem .8rem;}
      div[data-testid="stMetricValue"] {font-size:1.5rem; color:var(--mm-black);}
      div[data-testid="stMetricLabel"] {color:#666;}

      /* Primary buttons: black with yellow text (white on yellow is unreadable) */
      div.stButton > button[kind="primary"],
      div.stDownloadButton > button[kind="primary"] {
        background: var(--mm-black) !important; color: var(--mm-yellow) !important;
        border:1px solid var(--mm-black) !important; font-weight:700;
      }
      div.stButton > button[kind="primary"]:hover,
      div.stDownloadButton > button[kind="primary"]:hover {
        background:#000 !important; color:#fff !important;
      }
      div.stDownloadButton > button {border:1px solid var(--mm-black) !important;
        color: var(--mm-black) !important; font-weight:600;}

      button[data-baseweb="tab"][aria-selected="true"] {color: var(--mm-black) !important;}
      div[data-baseweb="tab-highlight"] {background-color: var(--mm-yellow) !important;}

      section[data-testid="stSidebar"] {background:#FAFAFA; border-right:1px solid var(--mm-line);}
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown(
    """
    <div class="mm-hero">
      <h1>📦 TikTok Stock Sync Tool</h1>
    </div>
    <div class="mm-bar">
      <b>Mister Mobile Singapore</b> &nbsp;·&nbsp; POS Masterlist → TikTok Seller Center bulk stock update
    </div>
    """,
    unsafe_allow_html=True,
)

DECISION_OPTIONS = [DECISION_PENDING, DECISION_LINK, DECISION_NOT_SELLING, DECISION_NOT_ON_TIKTOK]

for key, default in [("decisions", {}), ("filter_text", "")]:
    st.session_state.setdefault(key, default)


# --------------------------------------------------------------------------
# Sidebar — uploads
# --------------------------------------------------------------------------
with st.sidebar:
    st.header("1 · Upload files")

    pos_file = st.file_uploader("POS Masterlist (stock report)", type=["xlsx", "xlsm"])
    tt_file = st.file_uploader("TikTok Bulk Stock Inventory", type=["xlsx", "xlsm"])
    reg_file = st.file_uploader("SKU Registry (5 worksheets)", type=["xlsx", "xlsm"])

    with st.expander("Optional: seed mapping file"):
        st.caption(
            "First-time setup only — a 2-column sheet of "
            "**POS Stock Type ID → TikTok Variation/SKU ID**. "
            "The SKU Registry takes priority where both exist."
        )
        seed_file = st.file_uploader("POS ID ↔ TikTok Variation ID", type=["xlsx", "xlsm"],
                                     key="seed")

    st.divider()
    st.caption(
        "Required registry worksheets:\n\n"
        "· Locked Matches\n· New Masterlist SKUs\n· Match Review\n"
        "· Not Selling in TikTok\n· Not on TikTok Yet"
    )
    if st.button("🔄 Reset review decisions", use_container_width=True):
        st.session_state.decisions = {}
        st.rerun()


if not pos_file or not tt_file:
    st.info("⬅️ Upload the **POS Masterlist** and the **TikTok Bulk Stock Inventory** file to begin.")
    st.stop()


# --------------------------------------------------------------------------
# Parse
# --------------------------------------------------------------------------
@st.cache_data(show_spinner=False)
def _load(pos_bytes, tt_bytes, reg_bytes, seed_bytes):
    pos, pos_warn = parse_pos_masterlist(pos_bytes)
    tt_rows, meta = parse_tiktok_template(tt_bytes)
    reg = parse_registry(reg_bytes)
    seed = parse_seed_map(seed_bytes)
    reg.warnings.extend(pos_warn)
    return pos, tt_rows, meta, reg, seed


@st.cache_resource(show_spinner=False)
def _index(_rows, cache_key: str) -> MatchIndex:
    return MatchIndex(_rows)


pos_bytes = pos_file.getvalue()
tt_bytes = tt_file.getvalue()
reg_bytes = reg_file.getvalue() if reg_file else None
seed_bytes = seed_file.getvalue() if seed_file else None

try:
    with st.spinner("Reading workbooks…"):
        pos, tt_rows, meta, reg, seed = _load(pos_bytes, tt_bytes, reg_bytes, seed_bytes)
except SyncError as e:
    st.error(f"🛑 **Cannot continue** — {e}")
    st.stop()
except Exception as e:  # noqa: BLE001
    st.error(f"🛑 Unexpected error while reading files: {e}")
    st.stop()

match_index = _index(tt_rows, f"{len(tt_bytes)}:{len(tt_rows)}")
all_sku_ids = [r.sku_id for r in tt_rows if r.sku_id]
tt_sku_set = set(all_sku_ids)

if not reg_file and seed:
    st.info(
        f"No SKU Registry uploaded — running from the seed mapping file "
        f"({len(seed):,} links). Download the registry below and use it from the next run onward."
    )
elif not reg_file:
    st.warning(
        "No **SKU Registry** and no seed mapping uploaded. Every TikTok listing falls into "
        "Match Review and no quantities will change."
    )

# Seed prior decisions from the registry's New Masterlist SKUs sheet (once).
if reg.new_decisions and not st.session_state.decisions:
    prefill = {}
    for pid, (dec, link, note) in reg.new_decisions.items():
        canon = classify_decision(dec)
        if link and canon in (DECISION_LINK, DECISION_PENDING):
            prefill[pid] = {"decision": DECISION_LINK, "sku_id": link, "notes": note}
        elif canon in (DECISION_NOT_SELLING, DECISION_NOT_ON_TIKTOK):
            prefill[pid] = {"decision": canon, "sku_id": "", "notes": note}
    st.session_state.decisions = prefill


res = build_sync(pos, tt_rows, reg, seed, st.session_state.decisions)

# --------------------------------------------------------------------------
# Summary
# --------------------------------------------------------------------------
st.subheader("Validation summary")

c = st.columns(4)
c[0].metric("Locked Matches updated", len(res.locked_rows))
c[1].metric("Quantity cells changed", res.changed)
c[2].metric("New Masterlist SKUs", len(res.new_skus))
c[3].metric("SKUs requiring review", len(res.new_skus) + len(res.review_rows))

c = st.columns(4)
c[0].metric("Not Selling in TikTok", len(res.not_selling))
c[1].metric("Not on TikTok Yet", len(res.not_on_tiktok))
c[2].metric("Validation errors", len(res.errors))
c[3].metric("Unmatched records", len(res.review_rows))

st.caption(
    f"POS Masterlist: **{len(pos):,}** stock types · "
    f"TikTok inventory: **{len(tt_rows):,}** listing rows · "
    f"Quantity written to column **{meta['quantity_col']}** of the *Template* sheet."
)

if res.warnings:
    with st.expander(f"⚠️ Warnings ({len(res.warnings)})"):
        for w in res.warnings:
            st.write("·", w)

if res.zero_stock_notes:
    with st.expander(
        f"ℹ️ Locked IDs absent from today's POS report — counted as 0 ({len(res.zero_stock_notes)})"
    ):
        for n in res.zero_stock_notes[:200]:
            st.write("·", n)
        if len(res.zero_stock_notes) > 200:
            st.caption(f"…and {len(res.zero_stock_notes) - 200} more.")

if res.errors:
    with st.expander(f"🛑 Validation errors ({len(res.errors)})"):
        for e in res.errors[:300]:
            st.write("·", e)
        if len(res.errors) > 300:
            st.caption(f"…and {len(res.errors) - 300} more (see the Validation Errors sheet).")

st.divider()

# --------------------------------------------------------------------------
# Review workspace
# --------------------------------------------------------------------------
tab_review, tab_locked, tab_mr, tab_class = st.tabs(
    ["🔍 Review New Masterlist SKUs", "🔒 Locked Matches", "📋 Match Review", "🗂 Classified"]
)

with tab_review:
    total_new = len(res.new_skus)

    if total_new == 0:
        st.success("✅ No New Masterlist SKUs pending review. You can export.")
    else:
        st.write(
            f"**{total_new}** Masterlist SKU(s) with available stock are not yet matched, "
            "classified as *Not Selling in TikTok*, or *Not on TikTok yet*. "
            "All must be reviewed before export."
        )
        q = st.text_input("Filter by ID / brand / model / colour", key="filter_text")
        shown = [i for i in res.new_skus
                 if not q or q.lower() in f"{i.stock_id} {i.search_text}".lower()]

        page_size = 10
        pages = max(1, (len(shown) + page_size - 1) // page_size)
        page = st.number_input("Page", 1, pages, 1, key="review_page") if pages > 1 else 1
        chunk = shown[(page - 1) * page_size: page * page_size]

        taken = {sku for sku in
                 (d.get("sku_id") for d in st.session_state.decisions.values()) if sku}

        for item in chunk:
            cur = st.session_state.decisions.get(item.stock_id, {})
            with st.container(border=True):
                head = st.columns([3, 1])
                head[0].markdown(
                    f"**{item.label}**  \n"
                    f"<span class='mm-sku'>ID {item.stock_id} · {item.category or '—'}</span>",
                    unsafe_allow_html=True,
                )
                head[1].metric("Available", item.available)

                sugg = match_index.suggest(item, taken - {cur.get("sku_id", "")})
                options = [""] + [f"{r.sku_id} · {r.display[:80]}  ({sc:.0%})" for r, sc in sugg]
                sug_map = {options[i + 1]: sugg[i][0].sku_id for i in range(len(sugg))}

                cols = st.columns([2, 3, 2])
                prev = cur.get("decision", DECISION_PENDING)
                dec = cols[0].selectbox(
                    "Decision",
                    DECISION_OPTIONS,
                    index=DECISION_OPTIONS.index(prev) if prev in DECISION_OPTIONS else 0,
                    key=f"dec_{item.stock_id}",
                )
                pick = cols[1].selectbox(
                    "Suggested TikTok match (SKU ID · Seller SKU)",
                    options,
                    key=f"pick_{item.stock_id}",
                    help="Ranked against TikTok Seller SKU, Variation Option and Product name.",
                )
                manual = cols[2].text_input(
                    "…or paste SKU ID",
                    value=cur.get("sku_id", "") if not pick else "",
                    key=f"man_{item.stock_id}",
                )

                chosen = manual.strip() or sug_map.get(pick, "")
                if dec == DECISION_LINK and chosen and chosen not in tt_sku_set:
                    st.error(f"SKU ID `{chosen}` is not in the uploaded TikTok inventory file.")
                    chosen = ""
                if dec == DECISION_LINK and not chosen:
                    st.warning("Pick a suggestion or paste a SKU ID to complete this link.")

                if dec == DECISION_PENDING or (dec == DECISION_LINK and not chosen):
                    st.session_state.decisions.pop(item.stock_id, None)
                else:
                    st.session_state.decisions[item.stock_id] = {
                        "decision": dec,
                        "sku_id": chosen if dec == DECISION_LINK else "",
                        "notes": "",
                    }

                if not sugg:
                    st.caption("No close TikTok candidates found — likely *Not on TikTok yet*.")

        st.info(f"**{len(res.new_skus)}** still pending review.")
        st.button("Apply decisions / refresh", type="primary")

with tab_locked:
    st.write(
        f"**{len(res.locked_rows):,}** locked matches · "
        f"**{res.changed:,}** quantity cells will change."
    )
    only_changed = st.checkbox("Show only rows whose quantity changes", value=True)
    table = [
        {
            "SKU ID": d["sku_id"],
            "Seller SKU": d["seller_sku"],
            "Variation": d["variation_value"],
            "Masterlist ID(s)": d["masterlist_ids"],
            "Current Qty": d["current"],
            "New Qty": d["target"],
            "Δ": d["target"] - d["current"],
        }
        for d in res.locked_rows
        if not only_changed or d["target"] != d["current"]
    ]
    st.dataframe(table[:2000], use_container_width=True, hide_index=True)
    if len(table) > 2000:
        st.caption(f"Showing first 2,000 of {len(table):,} rows.")

with tab_mr:
    st.write(
        f"**{len(res.review_rows):,}** TikTok listings are not yet locked to a Masterlist ID. "
        "Their quantities are **left unchanged** in the export."
    )
    st.dataframe(
        [
            {
                "SKU ID": d["sku_id"],
                "Seller SKU": d["seller_sku"],
                "Variation": d["variation_value"],
                "Product": d["product_name"][:70],
                "Current Stock": d["current"],
            }
            for d in res.review_rows[:2000]
        ],
        use_container_width=True,
        hide_index=True,
    )

with tab_class:
    a, b = st.columns(2)
    with a:
        st.markdown(f"**Not Selling in TikTok** — {len(res.not_selling):,}")
        st.dataframe(
            [{"ID": i.stock_id, "Model": i.model, "Color": i.color, "Qty": i.available}
             for i in res.not_selling[:1500]],
            use_container_width=True, hide_index=True,
        )
    with b:
        st.markdown(f"**Not on TikTok Yet** — {len(res.not_on_tiktok):,}")
        st.dataframe(
            [{"ID": i.stock_id, "Model": i.model, "Color": i.color, "Qty": i.available}
             for i in res.not_on_tiktok[:1500]],
            use_container_width=True, hide_index=True,
        )

st.divider()

# --------------------------------------------------------------------------
# Export
# --------------------------------------------------------------------------
st.subheader("2 · Export")

blockers = []
if res.new_skus:
    blockers.append(f"{len(res.new_skus)} New Masterlist SKU(s) still need a review decision.")
if not reg_file and not seed:
    blockers.append("No SKU Registry or seed mapping uploaded — there is nothing to update.")
missing_sheets = [w for w in reg.warnings if "not found in the SKU file" in w]
if reg_file and missing_sheets:
    blockers.append("SKU Registry is missing required worksheet(s): "
                    + ", ".join(w.split("'")[1] for w in missing_sheets))

if blockers:
    st.error("🛑 **Export blocked**")
    for b in blockers:
        st.write("·", b)
else:
    st.success("✅ All checks passed — ready to export.")

stamp = datetime.now().strftime("%d-%m-%Y")
col_a, col_b = st.columns(2)

with col_a:
    if not blockers:
        tt_out = export_tiktok_template(tt_bytes, res.qty_updates, meta["quantity_col"])
        st.download_button(
            "⬇️ Download TikTok bulk upload file",
            data=tt_out,
            file_name=f"Tiktok_Stock_Update_{stamp}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )
        st.caption("Original template preserved — only the Quantity column is modified.")
    else:
        st.button("⬇️ Download TikTok bulk upload file", disabled=True, use_container_width=True)

with col_b:
    d = st.session_state.decisions.values()
    applied = (
        f"Applied from your review: "
        f"{sum(1 for x in d if x['decision'] == DECISION_LINK)} linked, "
        f"{sum(1 for x in d if x['decision'] == DECISION_NOT_SELLING)} not-selling, "
        f"{sum(1 for x in d if x['decision'] == DECISION_NOT_ON_TIKTOK)} not-on-tiktok."
    )
    reg_out = export_registry(res, pos, applied, sku_ids=all_sku_ids)
    st.download_button(
        "⬇️ Download updated SKU Registry",
        data=reg_out,
        file_name=f"Tiktok_Match_Review_UPDATED_{stamp}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    st.caption("Always available — save this and re-upload it next run.")
